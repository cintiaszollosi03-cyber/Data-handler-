import openpyxl
from openpyxl import Workbook
import os

from .. import generator
from ..model_dataclasses import Person, Workplace, Address

def write_people(people: list[Person],
                 workbook: openpyxl.Workbook,
                 sheet_name: str = "people",
                 heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "people")

    if heading:
        field_names = ["id", "name", "age", "male", "address"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)
        sheet.cell(row=row + offset, column=5, value=people[row].address.id if people[row].address else "")


def read_people(workbook: openpyxl.Workbook,
                sheet_name: str = "people") -> list[Person]:
    sheet = workbook[sheet_name]

    people = []
    row = 1
    
    if sheet.cell(row=1, column=1).value == "id":
        row = 2
    
    while True:
        if sheet.cell(row=row, column=1).value is None:
            break
        address_id = sheet.cell(row=row, column=5).value if sheet.max_column >= 5 else None
        person = Person(
            id=sheet.cell(row=row, column=1).value,
            name=sheet.cell(row=row, column=2).value,
            age=int(sheet.cell(row=row, column=3).value),
            male=bool(sheet.cell(row=row, column=4).value),
            address=address_id if address_id else None
        )
        row += 1
        people.append(person)
    return people


def write_workplaces(workplaces: list[Workplace],
                     workbook: openpyxl.Workbook,
                     sheet_name: str = "workplaces",
                     heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "workplaces")

    if heading:
        field_names = ["id", "name", "location", "employees"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(workplaces)):
        sheet.cell(row=row + offset, column=1, value=workplaces[row].id)
        sheet.cell(row=row + offset, column=2, value=workplaces[row].name)
        sheet.cell(row=row + offset, column=3, value=workplaces[row].location)
        employees_str = ",".join(workplaces[row].employees) if workplaces[row].employees else ""
        sheet.cell(row=row + offset, column=4, value=employees_str)


def read_workplaces(workbook: openpyxl.Workbook,
                    sheet_name: str = "workplaces") -> list[Workplace]:
    sheet = workbook[sheet_name]

    workplaces = []
    row = 1

    if sheet.cell(row=1, column=1).value == "id":
        row = 2
    
    while True:
        if sheet.cell(row=row, column=1).value is None:
            break
        
        employees_str = sheet.cell(row=row, column=4).value
        employees = employees_str.split(",") if employees_str else []
        employees = [emp.strip() for emp in employees if emp.strip()]
        
        workplace = Workplace(
            id=sheet.cell(row=row, column=1).value,
            name=sheet.cell(row=row, column=2).value,
            location=sheet.cell(row=row, column=3).value,
            employees=employees
        )
        row += 1
        workplaces.append(workplace)
    return workplaces


def write_addresses(addresses: list[Address],
                    workbook: openpyxl.Workbook,
                    sheet_name: str = "addresses",
                    heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "addresses")

    if heading:
        field_names = ["id", "street", "city", "country", "resident"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(addresses)):
        sheet.cell(row=row + offset, column=1, value=addresses[row].id)
        sheet.cell(row=row + offset, column=2, value=addresses[row].street)
        sheet.cell(row=row + offset, column=3, value=addresses[row].city)
        sheet.cell(row=row + offset, column=4, value=addresses[row].country)
        sheet.cell(row=row + offset, column=5, value=addresses[row].resident.id if addresses[row].resident else "")


def read_addresses(workbook: openpyxl.Workbook,
                   sheet_name: str = "addresses") -> list[Address]:
    sheet = workbook[sheet_name]

    addresses = []
    row = 1
    
    if sheet.cell(row=1, column=1).value == "id":
        row = 2
    
    while True:
        if sheet.cell(row=row, column=1).value is None:
            break
        
        resident_id = sheet.cell(row=row, column=5).value if sheet.max_column >= 5 else None
        
        address = Address(
            id=sheet.cell(row=row, column=1).value,
            street=sheet.cell(row=row, column=2).value,
            city=sheet.cell(row=row, column=3).value,
            country=sheet.cell(row=row, column=4).value,
            resident=resident_id if resident_id else None
        )
        row += 1
        addresses.append(address)
    return addresses


if __name__ == "__main__":
    import os
    from data.generator import generate_people, generate_workplaces, generate_addresses

    try:
        # Teszt könyvtár létrehozása
        test_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "data",
            "test_files"
        )
        os.makedirs(test_dir, exist_ok=True)

        # Tesztadatok generálása
        workplaces = generate_workplaces(2)
        addresses = generate_addresses(4)
        people = generate_people(4, workplaces.copy(), addresses.copy())

        # Új Excel munkafüzet létrehozása
        wb = openpyxl.Workbook()
        # Alapértelmezett üres lap eltávolítása
        wb.remove(wb.active)

        # Emberek tesztelése
        write_people(people, wb, "TesztEmberek")

        # Munkahelyek tesztelése
        write_workplaces(workplaces, wb, "TesztMunkahelyek")

        # Címek tesztelése
        write_addresses(addresses, wb, "TesztCimek")

        # Munkafüzet mentése
        xlsx_path = os.path.join(test_dir, "xlsx_teszt_adatok.xlsx")
        wb.save(xlsx_path)
        print(f"Munkafüzet elmentve ide: {xlsx_path}")

        # Beolvasás tesztelése
        wb_loaded = openpyxl.load_workbook(xlsx_path)

        loaded_people = read_people(wb_loaded, "TesztEmberek")
        loaded_workplaces = read_workplaces(wb_loaded, "TesztMunkahelyek")
        loaded_addresses = read_addresses(wb_loaded, "TesztCimek")

        print(f"Elmentve {len(people)} ember, beolvasva {len(loaded_people)} ember")
        print(f"Első beolvasott ember: {loaded_people[0] if loaded_people else 'Nincs adat'}")

        print(f"Elmentve {len(workplaces)} munkahely, beolvasva {len(loaded_workplaces)} munkahely")
        print(f"Első beolvasott munkahely: {loaded_workplaces[0] if loaded_workplaces else 'Nincs adat'}")

        print(f"Elmentve {len(addresses)} cím, beolvasva {len(loaded_addresses)} cím")
        print(f"Első beolvasott cím: {loaded_addresses[0] if loaded_addresses else 'Nincs adat'}")


    except ImportError:
        print("Az openpyxl nincs telepítve - XLSX kezelése nem lehetséges.")
        print("Telepíthető ezzel: pip install openpyxl")
